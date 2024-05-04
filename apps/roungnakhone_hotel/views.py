from django.http import Http404, HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Count
from django.db import transaction
from datetime import datetime
from django.utils import timezone
from django.template.loader import get_template
from openpyxl import Workbook
from weasyprint import HTML, CSS
import csv

from apps.account.forms import RegistrationForm
from .models import Receptionist, Guest, Room, Category, Reservation, Payment
from .forms import RoomForm, CategoryForm, ReservationForm, PaymentForm
from .decorators.authentication import guest_required, receptionist_required


def convert_to_vientiane_timezone(dt):
    if dt:
        local_tz = timezone.get_default_timezone()
        vientiane_offset = timezone.timedelta(hours=7)  # Asia/Vientiane is 7 hours ahead of UTC
        return dt - vientiane_offset
    else:
        return None


def convert_to_naive_datetime(dt):
    if dt:
        return dt.replace(tzinfo=None)
    else:
        return None


def get_authenticated_guest(user):
    if user.is_authenticated:
        if user.is_superuser or user.is_staff:
            return None  # Superusers and staff members don't have associated guests
        else:
            return Guest.objects.filter(user=user).first()
    else:
        return None


def get_authenticated_receptionist(user):
    if user.is_authenticated:
        if user.is_superuser or user.is_staff:
            return Receptionist.objects.filter(user=user).first()
        else:
            return None
    else:
        return None


#
# NOTE: Dashboard
#
@receptionist_required
def dashboard(request):
    # receptionist = None

    # if request.user.is_authenticated:
    #     receptionist = Receptionist.objects.filter(user=request.user).first()

    receptionist = request.receptionist

    context = {
        'receptionist': receptionist,
        'total_guests': Guest.total_guests(),
        'total_receptionists': Receptionist.total_receptionists(),
        'total_rooms': Room.total_rooms(),
        'available_rooms': Room.available_rooms(),
        'total_bookings': Reservation.total_bookings(),
        'total_revenue': Reservation.total_revenue(),
    }

    return render(request, 'dashboard/index.html', context)


@receptionist_required
def manage_receptionists(request):
    receptionist = Receptionist.objects.get(user=request.user)
    receptionists = Receptionist.objects.all()

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get('search', '')
    search_by = request.GET.get('search_by', 'username')  # Default to searching by username

    # Customize the query based on the search_by value
    if search_query and search_by == 'username':
        receptionists = Receptionist.objects.filter(user__username__icontains=search_query)
    elif search_query and search_by == 'email':
        receptionists = Receptionist.objects.filter(email__icontains=search_query)
    else:
        receptionists = Receptionist.objects.all().order_by('-updated_at')

    if request.method == 'POST':
        try:
            form = RegistrationForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                email = form.cleaned_data['email'].lower()
                user.is_staff = True
                user.save()
                new_receptionist = Receptionist.objects.create(
                    user=user,
                    email=email,
                )
                messages.success(request, 'Receptionist created successfully.')
            else:
                print('Form is not valid: ', form.errors)
        except Exception as e:
            print('ERROR:', str(e))

    context = {
        'receptionist': receptionist,
        'receptionists': receptionists,
    }
    return render(request, 'dashboard/pages/manage-receptionists.html', context)


@receptionist_required
def manage_guests(request):
    receptionist = Receptionist.objects.get(user=request.user)

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get('search', '')
    search_by = request.GET.get('search_by', 'username')  # Default to searching by username

    # Customize the query based on the search_by value
    if search_query and search_by == 'username':
        guests = Guest.objects.filter(user__username__icontains=search_query)
    elif search_query and search_by == 'email':
        guests = Guest.objects.filter(email__icontains=search_query)
    else:
        guests = Guest.objects.all().order_by('-updated_at')

    if request.method == 'POST':
        try:
            form = RegistrationForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                email = form.cleaned_data['email'].lower()
                user.is_staff = False
                user.save()
                new_guest = Guest.objects.create(
                    user=user,
                    email=email,
                )
                messages.success(request, 'Guest created successfully.')
            else:
                print('Form is not valid: ', form.errors)
        except Exception as e:
            print('ERROR:', str(e))

    context = {
        'receptionist': receptionist,
        'guests': guests,
    }
    return render(request, 'dashboard/pages/manage-guests.html', context)


@receptionist_required
def manage_rooms(request):
    rooms = Room.objects.all().order_by('-updated_at')
    categories = Category.objects.all().order_by('-updated_at')

    search_by = request.GET.get('search_by', 'number')
    search_query = request.GET.get('search', '')
    available = request.GET.get('available', False)
    category = request.GET.get('category', False)

    # Filter based on the search parameters
    if search_by == 'number':
        rooms = Room.objects.filter(number__icontains=search_query)
    else:
        # Handle other search criteria as needed
        pass

    # Additional filtering for availability and category
    if available:
        rooms = rooms.filter(available=True)

    if category:
        rooms = rooms.filter(category__category_name__icontains=category)

    if request.method == 'POST':
        room_form = RoomForm(request.POST, request.FILES)
        if room_form.is_valid():
            room_form.save()
            messages.success(request, 'Room created successfully.')
    else:
        room_form = RoomForm()

    context = {'rooms': rooms, 'categories': categories}

    return render(request, 'dashboard/pages/manage-rooms.html', context)


@receptionist_required
def manage_categories(request):
    categories = Category.objects.all().order_by('-updated_at')

    if request.method == 'POST':
        category_form = CategoryForm(request.POST)
        if category_form.is_valid():
            category_form.save()
            messages.success(request, 'Category created successfully.')
    else:
        category_form = CategoryForm()

    context = {'categories': categories}

    return render(request, 'dashboard/pages/manage-categories.html', context)


@receptionist_required
def manage_reservations(request):
    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get('search', '')
    search_by = request.GET.get('search_by', 'guest')

    # Customize the query based on the search_by value
    if search_query and search_by == 'guest':
        reservations = Reservation.objects.filter(guest__user__username__icontains=search_query)
    elif search_query and search_by == 'room':
        reservations = Reservation.objects.filter(room__number__icontains=search_query)
    else:
        reservations = Reservation.objects.all().order_by('-updated_at')

    if request.method == 'POST':
        reservation_form = ReservationForm(request.POST)
        if reservation_form.is_valid():
            reservation_form.save()
            messages.success(request, 'Reservation created successfully.')
    else:
        reservation_form = ReservationForm()

    # Convert check-in and check-out dates to Lao format
    # for reservation in reservations:
    #     reservation.check_in_date = reservation.check_in_date.strftime('%d/%m/%Y %I:%M %p')
    #     reservation.check_out_date = reservation.check_out_date.strftime('%d/%m/%Y %I:%M %p')

    context = {'reservations': reservations}

    return render(request, 'dashboard/pages/manage-reservations.html', context)


@receptionist_required
def manage_payments(request):
    payments = Payment.objects.all()

    search_query = request.GET.get('search', '')
    search_by = request.GET.get('search_by', 'payment_date')

    if search_query and search_by == 'payment_date':
        try:
            # Parse the datetime-local input
            date_search_query = datetime.strptime(search_query, '%Y-%m-%dT%H:%M')
            payments = Payment.objects.filter(payment_date__date=date_search_query)
        except ValueError:
            # Handle invalid datetime format gracefully
            payments = Payment.objects.none()
    else:
        # Handle the case when no search query is provided
        payments = Payment.objects.all()

    if request.method == 'POST':
        payment_form = PaymentForm(request.POST)
        if payment_form.is_valid():
            payment_form.save()
            messages.success(request, 'Payment created successfully.')
    else:
        payment_form = PaymentForm()

    context = {'payments': payments}

    return render(request, 'dashboard/pages/manage-payments.html', context)


def edit_room(request, pk):
    room = Room.objects.get(id=pk)
    if request.method == 'POST':
        room_form = RoomForm(request.POST, request.FILES, instance=room)
        if room_form.is_valid():
            room_form.save()
            messages.success(request, 'Room updated successfully.')
        else:
            messages.error(request, room_form.errors)
        return redirect('manage-rooms')
    else:
        room_form = RoomForm(instance=room)

    return render(request, 'dashboard/pages/manage-rooms.html')


def edit_category(request, pk):
    category = Category.objects.get(id=pk)
    if request.method == 'POST':
        category_form = CategoryForm(request.POST, instance=category)
        if category_form.is_valid():
            category_form.save()
            messages.success(request, 'Category updated successfully.')
        return redirect('manage-categories')
    else:
        category_form = CategoryForm(instance=category)

    return render(request, 'dashboard/pages/manage-categories.html')


def delete_receptionist(request, pk):
    receptionist = Receptionist.objects.get(pk=pk)

    if request.method == 'POST':
        receptionist.delete()
        return redirect('manage-receptionists')

    return render(request, 'dashboard/pages/manage-receptionist.html')


def delete_guest(request, pk):
    guest = Guest.objects.get(pk=pk)

    if request.method == 'POST':
        guest.delete()
        return redirect('manage-guests')

    return render(request, 'dashboard/pages/manage-guest.html')


def delete_room(request, pk):
    room = Room.objects.get(pk=pk)

    if request.method == 'POST':
        room.delete()
        return redirect('manage-rooms')

    return render(request, 'dashboard/pages/manage-rooms.html')


def delete_category(request, pk):
    category = Category.objects.get(pk=pk)

    if request.method == 'POST':
        category.delete()
        return redirect('manage-categories')

    return render(request, 'dashboard/pages/manage-categories.html')


#
# NOTE: Exports
#
def export_receptionists_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="receptionists.csv"'

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Last Name', 'Email', 'Phone'])

    receptionists = Receptionist.objects.all()
    for receptionist in receptionists:
        writer.writerow(
            [
                receptionist.first_name,
                receptionist.last_name,
                receptionist.email,
                receptionist.phone,
            ]
        )

    return response


def export_receptionists_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="receptionists.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Email', 'Phone'])

    receptionists = Receptionist.objects.all()
    for receptionist in receptionists:
        ws.append(
            [
                receptionist.first_name,
                receptionist.last_name,
                receptionist.email,
                receptionist.phone,
            ]
        )

    wb.save(response)
    return response


def export_guests_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="guests.csv"'

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Last Name', 'Email', 'Phone', 'Address'])

    guests = Guest.objects.all()
    for guest in guests:
        writer.writerow(
            [
                guest.first_name,
                guest.last_name,
                guest.email,
                guest.phone,
                guest.address,
            ]
        )

    return response


def export_guests_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="guests.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Email', 'Phone', 'Address'])

    guests = Guest.objects.all()
    for guest in guests:
        ws.append(
            [
                guest.first_name,
                guest.last_name,
                guest.email,
                guest.phone,
                guest.address,
            ]
        )

    wb.save(response)
    return response


def export_rooms_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rooms.csv"'

    writer = csv.writer(response)
    writer.writerow(['Room Number', 'Price', 'Category', 'Available'])

    rooms = Room.objects.all()
    for room in rooms:
        writer.writerow([room.number, room.price, room.category, room.available])

    return response


def export_rooms_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rooms.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['Room Number', 'Price', 'Category', 'Available'])

    rooms = Room.objects.all()
    for car in rooms:
        ws.append([car.number, car.price, car.category.category_name, car.available])

    wb.save(response)
    return response


def export_reservations_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reservations.csv"'

    writer = csv.writer(response)
    writer.writerow(['Guest', 'Room Number', 'Check-in Date', 'Check-out Date', 'Total Price'])

    reservations = Reservation.objects.all()
    for reservation in reservations:
        check_in_date = convert_to_vientiane_timezone(reservation.check_in_date)
        check_out_date = convert_to_vientiane_timezone(reservation.check_out_date)

        writer.writerow(
            [
                reservation.guest,
                reservation.room,
                check_in_date,
                check_out_date,
                reservation.total_price,
            ]
        )

    return response


def export_reservations_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reservations.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['Guest', 'Room Number', 'Check-in Date', 'Check-out Date', 'Total Price'])

    reservations = Reservation.objects.all()
    for reservation in reservations:
        check_in_date = convert_to_naive_datetime(
            convert_to_vientiane_timezone(reservation.check_in_date)
        )
        check_out_date = convert_to_naive_datetime(
            convert_to_vientiane_timezone(reservation.check_out_date)
        )

        ws.append(
            [
                reservation.guest.user.username,
                reservation.room.number,
                check_in_date,
                check_out_date,
                reservation.total_price,
            ]
        )

    wb.save(response)
    return response


def icon_font_awesome(request):
    return render(request, 'dashboard/pages/icon-fontawesome.html')


def error_page(request):
    return render(request, 'roungnakhone_hotel/pages/error-page.html')


#
# NOTE: Home -> Guests
#
@guest_required
def home(request):
    # guest = None

    # if request.user.is_authenticated:
    #     if request.user.is_superuser or request.user.is_staff:
    #         return redirect('dashboard')
    #     else:
    #         guest = Guest.objects.filter(user=request.user).first()

    guest = request.guest

    # Retrieve the four most reserved rooms
    top_reserved_rooms = Room.objects.annotate(num_reservations=Count('reservation')).order_by(
        '-num_reservations'
    )[:4]

    context = {
        'top_reserved_rooms': top_reserved_rooms,
        'guest': guest,
    }

    return render(request, 'roungnakhone_hotel/index.html', context)


@guest_required
def about(request):
    guest = None

    if request.user.is_authenticated:
        guest = Guest.objects.filter(user=request.user).first()

    context = {'guest': guest}

    return render(request, 'roungnakhone_hotel/pages/about.html', context)


@guest_required
def accomodation(request):
    guest = None

    if request.user.is_authenticated:
        guest = Guest.objects.filter(user=request.user).first()

    rooms = Room.objects.all()

    # Retrieve the four most reserved rooms
    top_reserved_rooms = Room.objects.annotate(num_reservations=Count('reservation')).order_by(
        '-num_reservations'
    )[:4]

    context = {'top_reserved_rooms': top_reserved_rooms, 'rooms': rooms, 'guest': guest}

    return render(request, 'roungnakhone_hotel/pages/accomodation.html', context)


@guest_required
def gallery(request):
    guest = None

    if request.user.is_authenticated:
        guest = Guest.objects.filter(user=request.user).first()

    context = {'guest': guest}

    return render(request, 'roungnakhone_hotel/pages/gallery.html', context)


@guest_required
def contact(request):
    guest = None

    if request.user.is_authenticated:
        guest = Guest.objects.filter(user=request.user).first()

    context = {'guest': guest}

    return render(request, 'roungnakhone_hotel/pages/contact.html', context)


#
# NOTE: Reservation and Payment
#
def reservation(request, pk):
    guest = None

    if request.user.is_authenticated:
        guest = Guest.objects.filter(user=request.user).first()

    room = Room.objects.get(id=pk)

    context = {'guest': guest, 'room': room}

    return render(request, 'roungnakhone_hotel/pages/reservation-form.html', context)


@transaction.atomic
def book_now(request):
    if request.method == 'POST':
        reservation_form = ReservationForm(request.POST)
        payment_form = PaymentForm(request.POST, request.FILES)

        if reservation_form.is_valid() and payment_form.is_valid():
            try:
                with transaction.atomic():
                    # Step 1: Save the Reservation instance
                    reservation = reservation_form.save()

                    # Step 2: Use the saved Reservation instance to create a new Payment instance
                    payment = payment_form.save(commit=False)
                    payment.reservation = reservation  # Link the Payment to the Reservation
                    payment.save()  # Step 3: Save the Payment instance

            except Exception as e:
                # Handle any exceptions that might occur during the transaction
                print(f"Error: {e}")
                return redirect('error-page')  # Redirect to an error page

            return redirect('payment-receipt', pk=payment.id)  # Redirect to a success page
        else:
            # If either form is not valid, you may want to handle this case
            # For example, you can display an error message or redirect to an error page
            # Print form errors for debugging
            print("Forms are not valid.")
            print("Reservation Form Errors:", reservation_form.errors)
            print("Payment Form Errors:", payment_form.errors)
            return redirect('error-page')
    else:
        reservation_form = ReservationForm()
        payment_form = PaymentForm()

    return redirect('home')


def payment_receipt(request, pk):
    try:
        payment = Payment.objects.get(id=pk)
    except Payment.DoesNotExist:
        raise Http404('Payment does not exist.')

    context = {'payment': payment}
    return render(request, 'roungnakhone_hotel/pages/payment-receipt.html', context)


def export_payment_to_pdf(request, pk):
    try:
        payment = Payment.objects.get(id=pk)
    except Payment.DoesNotExist:
        raise Http404('Payment does not exist.')

    template_path = 'roungnakhone_hotel/pages/payment-receipt.html'
    context = {'payment': payment}

    # Render the template
    template = get_template(template_path)
    html_content = template.render(context)

    # Set a dynamic filename based on rental information
    filename = f'payment_receipt_{payment.id}.pdf'

    # Create a WeasyPrint HTML object
    pdf_file = HTML(string=html_content).write_pdf(stylesheets=[CSS(string='@page { size: A4; }')])

    # Create a Django response with the PDF file
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="{filename}"'

    return response
